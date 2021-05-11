#!/usr/bin/env python3
# encoding : UTF-8
#%%
from copy import Error
from os import chdir, getcwd

from pandas._libs.tslibs import Timestamp

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import datetime
import pandas as pd
import win32com.client as win32
import os
import numpy as np
import csv
from datetime import datetime as Noseporque
import re
pd.options.mode.chained_assignment = None  # default='warn'

###
#TODO VER
#MENSAJE DE BIENVENIDA, EXPLICAR QUE LOS NOMBRES DE LOS REPORTES TIENEN QUE ESTAR EN MAYUSCULAS Y ESO
####################################################################################
################### PROTO MENSAJE ##################################
#####################################################################
while 1:
    try:
        msg = input("SMV/COV,LIRB/CIRB\n").split(',')
        visit_type = msg[0].upper()
        irb = msg[1].upper()
        if visit_type not in ["SMV","COV"] or irb not in ["LIRB","CIRB"]:
            raise Error 
        break
    except:
        print("Error, pone bien la info")
        continue
#%%
#Abro el archivo REPORT.xlsx para sacer la info necesaria
wb = openpyxl.load_workbook('REPORT.xlsx')    
ws=wb['Report']

#Extraigo el numero del sitio y el protocolo asi despues hago magia
Numero_de_sitio=int(ws["M2"].value)
Nombre_de_archivo=f'{str(Numero_de_sitio)} {visit_type} Site File Review.xlsx'
protocol = ws["K2"].value

#Esto me permite convertir la wea a .xlsx. lo saque de StackOverflow (NO CONSERVA H-LINKS)
#Hago lo mismo para los dos reportes de IP
if protocol == "RPC01-3101":
    prot_num = '3101'
else:
    prot_num = '3102'
    
while 1:
    if ((f'{prot_num} shipment'+'.xlsx') not in os.listdir('.')) and ((f'{prot_num} shipment'+'.xls') in os.listdir('.')):
        input("Por favor, para que el programa funcione el reporte de IP tiene que estar presente en la misma carpeta. Coloquelo y aprete enter")
    else:
        break
while 1:
    if (f'{prot_num} return'+'.xlsx') not in os.listdir('.'):
        input("Por favor, para que el programa funcione el reporte de IP devueltos tiene que estar presente en la misma carpeta. Coloquelo y aprete enter")
    else:
        break
while 1:
    if ('VISIT REPORT.csv'  in     os.listdir('.') and 
        'VISIT REPORT.xlsx' not in os.listdir('.')):
        input("Por favor, para que el programa funcione el reporte visitas tiene que estar presente en la misma carpeta. Coloquelo y aprete enter")
    else:
        break



# Voy copiando las columnas que necesito
Excel={}
Encabezados=['Level','Category','Document Type','Ref Model Subtype','Study Item Name','Ref Model ID','Site Personnel Name','Document Date','Expiration Date']
Max_Column=ws.max_column

# Busco usando la lista de encabezados, las coordenadas
for Encabezados_de_colunmnas in Encabezados:
    Excel[Encabezados_de_colunmnas]=[]
    for col_num in range (1,Max_Column):
        if ws.cell(row=1,column=col_num).value==Encabezados_de_colunmnas:            
            Letra_de_columna=get_column_letter(col_num)
            if ws.cell(row=1,column=col_num).value=='Document Date' or ws.cell(row=1,column=col_num).value=='Expiration Date':
                for cell in ws[Letra_de_columna]:
                        try: #los que son "None"
                            cell.value=cell.value.strftime('%d-%b-%Y')
                        except:
                            continue
            for cell in ws[Letra_de_columna]:
                Excel[Encabezados_de_colunmnas].append(cell.value)
# Con esto ya saque toda la info del REPORT.xlsx y lo tengo en el diccionario "Excel"

#Abro el archivo TEMPLATE.xlsx para pegar la info sacada de REPORT.xlsx.

wb = openpyxl.load_workbook('TEMPLATE.xlsx')    
ws=wb['Site']
Max_Row=ws.max_row
Max_Column=ws.max_column

# Antes de cualquier cosa, los encabezados del template y del report no son iguales, asi que hago la conversion asi no hay problema

Excel['Class']=Excel['Level']
Excel['Document Category']=Excel['Category']
Excel['Document Name']=Excel['Study Item Name']
Excel['Site personnel name']=Excel['Site Personnel Name']
Excel['Document date']=Excel['Document Date']
Excel['Expiration date']=Excel['Expiration Date']
# Una vez esto hecho, a agregar el contenido de esas listas al fondo de cada columna
# Parseo el diccionario vs los encabezados

for key in Excel.keys():
    for column_num in range (1,Max_Column): #Reviso todos los encabezados
        if key==ws.cell(1,column_num).value: #Si el encabezado coincide con la key que tengo
            for row_num in range (1,len(Excel[key])): 
                ws.cell(column=column_num,row=Max_Row+row_num-2).value=Excel[key][row_num] #Agrego todos los valores de la lista al final del ws

# Guardo todo antes del procesador

wb.save(Nombre_de_archivo)
#%%

#Abro el documento con Pandas

filename=os.getcwd()+'\\'+Nombre_de_archivo #consigo la direccion del archivo. con el os.getcwd() obtengo la dir del directorio donde esta el programa

SFR= pd.read_excel(filename, sheet_name='Site',header=0)
wb = openpyxl.load_workbook(Nombre_de_archivo)    
ws=wb['Site']
#Pongo las dates en formato datetime
def Str_to_date(str):
    try:
        return Noseporque.strptime(str,'%d-%b-%Y').date()
    except:
        return None
    
for index, row in SFR['Document date'].iteritems():
    SFR.loc[index,'Document date']=Str_to_date(SFR.loc[index,'Document date'])
    SFR.loc[index,'Expiration date']=Str_to_date(SFR.loc[index,'Expiration date'])

#Creo este objeto para poder almacenar la info
class Sitio:
    def add_atribute(atribute, info_to_add):
        '''Esta funcion sirve para poder crear un atribute y asignarle un value llamado "info_to_add". Esto es especialmente util si estoy agregando cosas en un for loop'''
        if type(atribute)==str:
            atribute=atribute.replace(' ','_') #depuro los posibles caracteres q puedan joder al ponerle el atributo al objeto. TODO usar regex
            atribute=atribute.replace('/','_')
        if hasattr(Sitio,atribute):
            New_info=getattr(Sitio,atribute)
            New_info.append(info_to_add) #Siexiste el atributo, lo appendeo. Ya que siempre es una lista
        else:
            New_info=[info_to_add] #else, lo hago una lista.
        setattr(Sitio,atribute,New_info)

    #Informacion del Sitio
    Site_Number = Numero_de_sitio
    Cerrado = False
    #IP Shipment information
    First_IP=''    
    IP_Recieved=[]
    IP_Returned=[]
    

#%%
#usando el reporte de visitas, checkear que estan todas las cfm, fup, svr
    

#Agarro del Visit report y agrego al Sitio cada una de las visitas, usando como nombre de atributo el tipo de visita, y el atributo es la fecha..
#El atributo es una lista, y si hay mas de una visita del mismo tipo, lo appendea
#Primero leo el visit report y filtro por el sitio
Visit_Report= pd.read_excel( os.getcwd()+"\\VISIT.xlsx",header=2)
Visit_Report = Visit_Report.loc[(Visit_Report["Protocol #"] == protocol) & (Visit_Report["Site #"] == Numero_de_sitio)]
for index_Visit_Report,row_Visit_Report in Visit_Report['Visit Type'].iteritems():    
    try:
        Sitio.add_atribute(Visit_Report['Visit Type'][index_Visit_Report], Visit_Report['Visit End'][index_Visit_Report])
    except:
        pass
#Ahora parseo por el DF del excel

def add_to_excel(Row_num,Ref_model_ID,Present_in_eTMF,Comments,Action_needed,*Action, **staff):
    '''
    Esta Funcion sirve para agregar los comentarios al Excel. 
    Row_num = La fila donde va a ir el comentario
    Ref_model_ID = El codigo del eTMF correspondiente al archivo
    Present_in_eTMF = Si esta presente o no
    Coments = Que comentarios van en la columna "Coments"
    Action_needed = Y/N, si es necesario hacer algo
    *Action = Si es necesario hacer algo, esto dice que
    **Staff = El staff member al que corresponde este documento, este o no. Esto corresponde a la columna "G": Site personnel name
    
    
    '''
    wb = openpyxl.load_workbook(Nombre_de_archivo)    
    ws=wb['Site']
    if Present_in_eTMF=='N':
        Row_num=ws.max_row+1 #Si no esta presente, mando el comentario al fondo
    else:
        Row_num +=2 #para la df el primer index es 0, pero el excel arranca en 2
    if staff:
        for arg in staff.values():
            ws.cell(Row_num,7).value = arg   
    ws.cell(Row_num,6).value  = Ref_model_ID
    ws.cell(Row_num,11).value = Present_in_eTMF
    ws.cell(Row_num,12).value = Comments
    ws.cell(Row_num,13).value = Action_needed
    if Action_needed == 'Y':
        ws.cell(Row_num,14).value=Action[0]
    wb.save(Nombre_de_archivo)

def add_visit_from_report(Ref_ID, visit_date):
    '''Esta funcion checkea todas las visitas del reporte de visitas y se fija si estan en el archivo de SFR. Si estan, escribe en comments 
    diciendo que es el archivo, si no esta agrega al final una linea con la info de que es lo que falta
    Ref_ID : Codigo de eTMF del tipo de visita a introducir
    visit_date : Fecha de la visita, en Timestamp o datetime
    
    
    
    '''
    if visit_date > datetime.datetime.today():
        return None
    df = SFR.loc[SFR["Ref Model ID"] == Ref_ID]
    Letter_Types=['Confirmation Letter','Follow-up Letter', 'Monitoring Report']
    #TODO PARA las confirmation letter, usar tmb "planned date", porq a veces si se postearga la visita la confirmation letter te queda en cualquiera
    margen_de_error = 14
    margen_de_error = str(margen_de_error) + ' days'
    for index, row in df['Ref Model ID'].iteritems():
        fecha_documento = pd.to_datetime(df.loc[index,'Document date'] )
        subtipo_de_documento = df.loc[index,'Ref Model Subtype']
        codigo_del_documento = df.loc[index,'Ref Model ID']
        if fecha_documento == None:
            continue
        if fecha_documento == visit_date:
            diferencial = pd.Timedelta(0)
        else:
            if subtipo_de_documento == 'Confirmation Letter':
                diferencial = visit_date - fecha_documento
            else:
                diferencial = fecha_documento - visit_date
        if  (diferencial == pd.Timedelta(0)) or (abs(diferencial) < pd.Timedelta(margen_de_error)) and (codigo_del_documento==Ref_ID):               
            if subtipo_de_documento not in Letter_Types:
                if subtipo_de_documento not in ['Confirmation Letter','Follow-up Letter', 'Monitoring Report']:
                    add_to_excel(index,Ref_ID,'Y',"Please check this document",'Y','Please check this document')
                else:
                    add_to_excel(index,Ref_ID,'Y',f"Duplicated {(subtipo_de_documento)} from {(visit_date.date())} visit",'Y','Errase Duplicated')
                SFR.drop(index,inplace=True)
                continue #Si tengo un duplicado, no va a estar en letter types xq ya fue popeado. 
            else:
                Letter_Types.remove(subtipo_de_documento)
                SFR.drop(index,inplace=True)
                add_to_excel(index,Ref_ID,'Y',f"{(subtipo_de_documento)} from {(visit_date.date())} visit",'N')
    if Letter_Types!=[]:
        add_to_excel(0,Ref_ID,'N',f'{Letter_Types} missing from {(visit_date.date())} visit','Y','Collect from Site') #el primer argumento no importa en este caso, ya que se va a a setear igual al fondo

def check_and_add(code, atribute):
    '''Esta Funcion agarra un Ref ID y se fija si en el objeto Sitio tengo un tipo de visita que corresponda a ese ID. Si esta, ejecuta add_visit_from_report'''
    if hasattr(Sitio,atribute):
        for visit_date in getattr(Sitio,atribute):
            add_visit_from_report(code, visit_date)
        if atribute == 'Site_Visit_Closeout' or atribute == 'Telephone_Closeout':
            Sitio.Cerrado = True
#%%

check_and_add('05.01.04','Site_Visit_Selection')
check_and_add('05.03.01','Site_Visit_Initiation')            
check_and_add('05.04.03','Site_Visit_Interim')  
check_and_add('05.04.08','Site_Visit_Closeout' )
check_and_add('05.04.08','Telephone_Closeout' )
check_and_add('05.04.05','Training/Booster Visit' )

#%%
#######################################################################################
############################### IP SHIPMENT  ##########################################
#######################################################################################
#TODO poner todo lo que falta en una misma linea asi no es un sida
#Extraigo informacion de IP SHIPMENT y lo meto en el Sitio. 
if protocol == "RPC01-3101":
    header = 1
else:
    header = 2
    

IP_SHIPMENT= pd.read_excel((f'{prot_num} shipment'+'.xlsx'), sheet_name='Sheet',header=header)

#Reduzco a mi sitio y a los envios recibidos
IP_SHIPMENT=IP_SHIPMENT.loc[IP_SHIPMENT['Shipment Status']=='Received']
IP_SHIPMENT_Site=IP_SHIPMENT.loc[IP_SHIPMENT['Ship to Site Number']==int(Sitio.Site_Number)]

if IP_SHIPMENT_Site.empty: #Puede que no haya tenido IP el sitio
    Sitio.IP_Recieved=None
    Sitio.First_IP=None
else:
    IP_Shipping_Dates=pd.to_datetime(IP_SHIPMENT_Site['Received Date'])

#Guardo todos los envios y el primero, porque me sirve para los IP temperature logs    
    Sitio.IP_Recieved=list(IP_SHIPMENT_Site['Shipment Number'])
    Sitio.First_IP=min(IP_Shipping_Dates).strftime('%d-%b-%Y')

#Ahora busco en SFR si estan los IP shipments
SFR_test=SFR.loc[SFR['Ref Model ID']=='06.01.04']
if Sitio.IP_Recieved != None:
    for shipment in Sitio.IP_Recieved:
        Shipment_types=['Packing order','Shipment confirmation','Acknowledgement of Receipt']
        Bacon=SFR_test.loc[SFR_test['Document Name'].str.contains(str(shipment), flags=re.IGNORECASE,na=False)]
        for documents in Shipment_types:
            if Bacon.index[Bacon['Document Name'].str.contains(documents, flags=re.IGNORECASE,na=False, regex=True)].empty==False:
                spam=Bacon.index[Bacon['Document Name'].str.contains(documents)]
                Shipment_types.remove(documents)
                if documents=='Acknowledgement':
                    add_to_excel(spam[0],'06.01.04','Y',"Check if this file is a Packing List, Shipping confirmation or Shipping Request",'N')
                else:
                    try:
                        add_to_excel(spam[0],'06.01.04','Y',f"{documents} for {shipment} shipping",'N')
                    except IndexError as e:
                        print(f"Error al procesar {documents} for {shipment} shipping.")
                        print(f"Motivo: {e}")
                        pass
        if Shipment_types!=[]:
            if 'Acknowledgement of Receipt' in Shipment_types:
                Shipment_types.remove('Acknowledgement of Receipt')
            if len(Shipment_types) == 3:
                msg = 'IP Packing list and IP Shipment confirmation'
            else:
                msg = ', '.join(Shipment_types)
            add_to_excel(0,'06.01.04','N',f'{msg} missing for {shipment} shipment','Y','Collect from Site') #el primer argumento no importa en este caso, ya que se va a a setear igual al fondo
    #TODO this needs more work. no funciona bien, me da cualquier cosa

    #Extraigo informacion de IP RETURN y lo meto en el Sitio
    IP_RETURN= pd.read_excel((f'{prot_num} return'+'.xlsx'), sheet_name='Sheet',header=header)
    IP_RETURN=IP_RETURN.loc[IP_RETURN['Return Shipment Status']=='Received']
    IP_RETURN_Site=IP_RETURN.loc[IP_RETURN['Ship from Site Number']==int(Sitio.Site_Number)]

    if IP_RETURN_Site.empty: #Puede que no haya devuelto IP el sitio
        Sitio.IP_Returned=None  
    else:
        IP_Return_Dates=pd.to_datetime(IP_RETURN_Site['Date Received'])

    #Guardo los IP return
        Sitio.IP_Returned=list(IP_RETURN_Site['Return Shipment Number'])
    if Sitio.IP_Returned != None:
        SFR_test=SFR.loc[SFR['Ref Model ID']=='06.01.10']
        if Sitio.IP_Returned != None:
            for shipment in Sitio.IP_Returned:    
                Bacon=SFR_test.loc[SFR_test['Document Name'].str.contains(str(shipment), flags=re.IGNORECASE,na=False)]
                if Bacon.index[Bacon['Document Name'].str.contains(str(shipment), flags=re.I, regex=True)].empty==False:
                    spam=Bacon.index[Bacon['Document Name'].str.contains(str(shipment), flags=re.I, regex=True)]
                    add_to_excel(spam[0],'06.01.10','Y',f"IP Return documentation for {shipment} shipping",'N')
                else:
                    add_to_excel(0,'06.01.10','N',f"Missing IP Return Documentation for {str(shipment)} shipping",'Y','Collect from site')
        else:    
            add_to_excel(SFR.index[SFR['Ref Model ID'] == '06.01.10'][0],'06.01.10','Y',"No IP was returned",'N')
    #Usando el primer Ip shipment, defino desde cuando necesito los IP temp logs y calibration logs
    
    add_to_excel(0,'06.04.01','N',f"Please check that the IP temperature logs are present from {Sitio.First_IP} to present.",'Y','Collect from site, if applicable')
    add_to_excel(0,'06.04.03','N',f"Please check that the calibration logs are present from {Sitio.First_IP} to present.",'Y','Collect from site, if applicable')
else:
    add_to_excel(0,"06.01.04","N","Site never had IP. Shipping documentation, IP temperature log, IP accountability logs, etc are not applicable","N")
#%%
#################################################################################################################################
################ LOGS ###########################################################################################################
################################################################################################################################

if visit_type == "SMV":
    logs = "Updated DOA, subject log (will need it even if screening failed), Updated IP accountability and Protocol Deviation log (if applicable)  and updated visit log."
else:
    logs = "Final DOA, subject log (will need it even if screening failed), Final IP accountability log and Protocol Deviation log (if applicable) , Final visit log with PI signature"
add_to_excel(0,'05.02.18','N',logs,'Y','Collect from site')
    
#%%
#usar un reporte de CTMS para predecir el study team (PIs, SubIs).
while 1:
    if (f"CONTACT {prot_num}.csv") in os.listdir('.') and (f"CONTACT {prot_num}.xlsx") not in os.listdir('.'):
        input(f"Por favor introduzca el reporte de CONTACT {prot_num}.xlsx en el mismo directorio del programa")
    else:
        break
Contact_Report= pd.read_excel(os.getcwd()+f"\\CONTACT {prot_num}.xlsx",header=0)
SFR= pd.read_excel(filename, sheet_name='Site',header=0)
Contact_Report= Contact_Report.loc[(Contact_Report["Site #"] == Numero_de_sitio)]
Contact_Report=Contact_Report[['Role','First Name', 'Last Name', 'Start Date','End Date']]

class Site_Staff:
    """
    This class holds the information of the staff members

    Atributes:
        name (str): Name of the person.
        last_name (str): Last name of the person.
        role (str): Role of the person.
        start_date(timestamp): Starting date for that person on the site.
        end_date(timestamp): End date for that person on the site. This might not be present if the site member is still active.
        
    """

    def __init__ (self, name=None, last_name=None, role=None, start_date=None,end_date=None):
        """
        Constructor of Site_staff class.

        Parameters:
            name (str): Name of the person.
            last_name (str): Last name of the person.
            role (str): Role of the person.
            start_date(timestamp): Starting date for that person on the site.
            end_date(timestamp): End date for that person on the site. This might not be present if the site member is still active.
        """
        self.name = name
        self.last_name = last_name
        self.role= role
        self.GCP = False
        self.EDC = False
        self.IATA = False
        self.License = False
        if self.role == 'Principal Investigator':
            self.GCP = True
            self.EDC = True
            self.IATA = True
            self.License = True
        elif self.role == 'Sub-Investigator':
            self.GCP = True
            self.License = True
        if type(start_date) == str:
            start_date = pd.Timestamp(start_date)
        self.start_date = start_date
        if pd.isna(end_date):
            self.end_date = 'Present'
        else:
            if type(end_date) == str:
                end_date =  pd.Timestamp(end_date)
            self.end_date = end_date

Names= Contact_Report['First Name'].tolist()
Last_names=Contact_Report['Last Name'].tolist()
Roles=Contact_Report['Role'].tolist()
Starting_dates=Contact_Report['Start Date'].tolist()
Ending_dates=Contact_Report['End Date'].tolist()

Site_Members = []

for nombres,apellidos,funcion,inicio,fin in zip(Names,Last_names,Roles,Starting_dates,Ending_dates):
    Site_Members.append(Site_Staff(nombres,apellidos,funcion,inicio,fin))

Sitio.Site_members=Site_Members
#%%
#Una vez que tengo la informacion guardada la uso para que haga cosas

pd.options.mode.chained_assignment = None  # default='warn'

SFR= pd.read_excel(filename, sheet_name='Site',header=0)
SFR['Document date']=pd.to_datetime(SFR['Document date'])
SFR['Expiration date']=pd.to_datetime(SFR['Expiration date'])

#Armo una df con lo q me interesa bc reasons
SFR_trainings= SFR.loc[(SFR['Ref Model ID'] == '05.02.07') | (SFR['Ref Model ID'] == '05.03.03')]

#Planteo los posibles certificados, previamente definidos en la clase
Certificates = ['GCP', 'EDC', 'IATA', 'License']

#Parseo por todos los staff members
for staff_member in Sitio.Site_members:    
    #Reduzco la df a solo lo que tiene el apellido del staff member en el nombre del archivo o en la columna de "site personnel name" (esta ultima a veces esta vacia xq la mtadata es malisima)
    df = SFR_trainings.loc[(SFR_trainings['Site personnel name'].str.contains(staff_member.last_name,na=False, flags=re.I, regex=True)) | (SFR_trainings['Document Name'].str.contains(staff_member.last_name,na=False, flags=re.I, regex=True))]
  
    #Por cada atributo en Certificates...    
    for atribute in Certificates: 
               
        #Si el atributo es True
        if getattr(staff_member,atribute) == True:
            
            #Me fijo si hay algun archivo presente que tenga el nombre del certificado (atribute) en el subtype o en el nombre
            #Si no hay nada, agrego una columna al final pidiendo lo que falta
            #Para evitar codigo feo, defino una nueva df_cert para no estar typeando df.loc[(df['Ref Model Subtype'].str.contains(atribute)) | (df['Document Name'].str.contains(atribute))]
            #todo el tiempo
            df_cert = df.loc[(df['Ref Model Subtype'].str.contains(atribute, flags=re.I, regex=True)) | (df['Document Name'].str.contains(atribute, flags=re.I, regex=True))]                
            if atribute == 'GCP' or atribute == 'License':
            #TODO hay un bug q cuando la licencia tiene como date antes q otra pero aplica hasta despues, se pornea todo. ver bien como arreglar
                Ref_model= '05.02.07'
            else:
                Ref_model= '05.03.03'
                
            #Si no encuentro resultados, agregar al fondo
            if staff_member.end_date == 'Present':
                msg = 'Present'
            elif type(staff_member.end_date) == str:
                msg = datetime.datetime.strptime(staff_member.end_date,"%d-%b-%Y").date()
            if type(staff_member.start_date) == str:
                staff_member.start_date = datetime.datetime.strptime(staff_member.start_date,"%d-%b-%Y")

            if df_cert.empty:                
                add_to_excel(' ',Ref_model, 'N', f'{atribute} for {staff_member.last_name} covering from {staff_member.start_date.date()} to {msg}', 'Y', 'Collect from site')
                  
            #Si encontro archivos vamos a checkear la fecha y compararla con lo que se necesita
            else:
                #ordeno la DF por fecha creciente
                df_cert.sort_values(by='Document date', inplace=True)
                df_cert.reset_index(inplace=True)
                
                #evaluo todos los items en la df
                #Para ir checkeando necesito ir trackeando las fechas cubiertas. Para esto creo                
                Cert_date = staff_member.start_date
                
                #ahora parseo por toda la df en orden creciente                
                for index in df_cert.index:  
                 #Como algunas certificaciones no tienen exp date porque la metadata es un sida, lo arreglo aca
                    if atribute == 'GCP':
                        df_cert['Expiration date'][index] = (df_cert['Document date'][index]+datetime.timedelta(days=1095)).date()
                    elif atribute == 'EDC':
                        df_cert['Expiration date'][index] = 'End of study'
                    elif atribute == 'IATA': 
                        df_cert['Expiration date'][index] = (df_cert['Document date'][index]+datetime.timedelta(days=730)).date()
                    elif atribute == 'License' and pd.isna(df_cert.loc[index,'Expiration date']):
                        df_cert['Expiration date'][index] = (df_cert['Document date'][index]+datetime.timedelta(days=365)).date()             
                    
                    #Ahora extraigo el index correspondiente al row en la SFR original y agrego la info en los comments
                    comment = f"{atribute} certificate from {df_cert['Document date'][index].date()} to {df_cert['Expiration date'][index]}"
                    add_to_excel(df_cert['index'][index],Ref_model,'Y',comment , 'N', staff=f'{staff_member.name} {staff_member.last_name}')  

                    #si la diferencia de fecha entre la licencia/training y la fecha de inicio/licencia anterior es mayor a 0, significa q el training ocurrio antes de la fecha limite, ergo esta todo bien
                    #Pero si es menor a 0, significa q el certificado se expidio despues de la fecha limite.
                    #seteo unos 90 de gracia para que la dif este todo bien, pero si es mayor a esos 90 dias hago cosas
                    if atribute != 'EDC':
                        if type(Cert_date) == str:
                            Cert_date = datetime.datetime.strptime(Cert_date,'%d-%b-%Y')
                        if (df_cert['Document date'][index] - Cert_date) > datetime.timedelta(days=90):                             
                            add_to_excel(df_cert['index'][index],Ref_model,'N',f"Missing {atribute} certificate for {staff_member.last_name}, {staff_member.name} covering from {Cert_date.date()} to {df_cert['Document date'][index].date()}", 'Y', 'Collect from site',staff=(f'{staff_member.name} {staff_member.last_name}'))
                        Cert_date = pd.to_datetime(df_cert['Expiration date'][index]          )
                    
                #checkeo la dif entre cuando vence la ultima licencia y cuando se fue del sitio o presente
                if atribute != 'EDC':
                    if staff_member.end_date == 'Present':
                         if (pd.Timestamp(datetime.datetime.today().date()) - pd.Timestamp(Cert_date)) > datetime.timedelta(days=0):
                            add_to_excel(0,Ref_model, 'N', f'Missing {atribute} certificate for {staff_member.last_name}, {staff_member.name} from {Cert_date.date()} to {msg}.', 'Y', 'Collect from site, if applicable', staff=f'{staff_member.name} {staff_member.last_name}')
                    else:
                        if type(staff_member.end_date) == str:
                            msg = pd.Timestamp(staff_member.end_date)
                        else:
                            msg = staff_member.end_date
                        if (msg - Cert_date) > datetime.timedelta(days=0):
                            add_to_excel(0,Ref_model, 'N', f'Missing {atribute} certificate for {staff_member.last_name}, {staff_member.name} from {Cert_date.date()} to {msg}.', 'Y', 'Collect from site, if applicable', staff=f'{staff_member.name} {staff_member.last_name}')
#%%
#CVs
SFR = pd.read_excel(filename, sheet_name='Site',header=0)
cv = SFR.loc[(((SFR["Ref Model ID"] == "05.02.04") | (SFR["Ref Model ID"] == "05.02.05")) | (SFR["Ref Model ID"] == "05.02.06")) & (~(SFR["Document Name"]).isna())]      
lista_pi_y_subi = [(investigator.last_name, investigator.name, investigator.role)  for investigator in Sitio.Site_members if (investigator.role == 'Principal Investigator' or  investigator.role == 'Sub-Investigator')]
for investigator in lista_pi_y_subi:    
    if cv.loc[cv["Document Name"].str.contains(investigator[0],regex=True,flags=re.IGNORECASE)].empty:
        if investigator[2] == 'Principal Investigator':
            codigo = "05.02.04"
        elif  investigator[2] == 'Sub-Investigator':
            codigo = "05.02.05"
        add_to_excel(0,codigo, "N", f"Missing {investigator[0]}, {investigator[1]} CV", "Y", "Collect from site", f"{investigator[0]}, {investigator[1]}")
    else:
        indice = cv.loc[cv["Document Name"].str.contains(investigator[0],regex=True,flags=re.IGNORECASE)].index.values[0]
        add_to_excel(indice,(cv["Ref Model ID"][indice]), "Y", f"{investigator[0]}, {investigator[1]} CV", "N")
#%%
#Data Privacy, basicamente lo mismo que los CVs, pero con la condicion que start date > nov2018
SFR = pd.read_excel(filename, sheet_name='Site',header=0)
dp = SFR.loc[(SFR["Ref Model ID"] == "05.02.11") & (~(SFR["Document Name"]).isna())]      
lista_pi_y_subi = [(investigator.last_name, investigator.name, pd.Timestamp(investigator.start_date), investigator.end_date)  for investigator in Sitio.Site_members if (investigator.role == 'Principal Investigator' or  investigator.role == 'Sub-Investigator')]

for investigator in lista_pi_y_subi:    
    if (investigator[3] == "Present") or  (pd.Timestamp ("2019-01-01") >= investigator[3]):
        if dp.loc[dp["Document Name"].str.contains(investigator[0],regex=True,flags=re.IGNORECASE)].empty:
            add_to_excel(0,"05.02.11", "N", f"Missing Data Privacy Form for {investigator[0]}, {investigator[1]}", "Y", "Collect from site", f"{investigator[0]}, {investigator[1]}")
        else:
            indice = dp.loc[dp["Document Name"].str.contains(investigator[0],regex=True,flags=re.IGNORECASE)].index.values[0]
            add_to_excel(indice,(dp["Ref Model ID"][indice]), "Y", f"Data Privacy for {investigator[0]}, {investigator[1]}", "N")

#%%
#FDFs. No se me ocurre otra manera de hacerlo q repetir el codigo
# FDF (3 kinds) - Receptos (start, as of Sept 2013), Celgene (effective Mar2016), BMS (effect Jan2020)
#primero agrego una columna llamada version, asi puedo diferenciar que FDF es
fdf = SFR.loc[(SFR["Ref Model ID"] == "05.02.10")& (~SFR["Document Name"].isna())]
fdf["Version"] = pd.Series(dtype="object")
for index,row in fdf.iterrows():
    try:
        document_date = pd.Timestamp(fdf["Document date"][index])
        if (document_date > pd.Timestamp("2013-09-01")) and (document_date < pd.Timestamp("2016-03-01")):
            version = "Receptos"
        elif (document_date > pd.Timestamp("2016-03-01")) and (document_date < pd.Timestamp("2020-01-01")):
            version = "Celgene"
        else:
            version = "BMS"
        fdf["Version"][index] = version
    except:
        fdf["Version"][index] = "Check this document"

#ahora, dependiendo de cuando arranco el investigador, y cuando termino puedo predecir que FDFs necesita, y me fijo si estan
for investigator in lista_pi_y_subi:    
    #filtro las FDF por el investigator
    inv_fdf = fdf.loc[fdf["Site personnel name"].str.contains(investigator[0], regex=True,flags=re.IGNORECASE)]
    if investigator[2]< pd.Timestamp("2016-03-01"):
        if inv_fdf.loc[inv_fdf["Version"] == "Receptos"].empty:
            add_to_excel(0,"05.02.10", "N", f"Missing Receptos FDF for {investigator[0]}, {investigator[1]}", "Collect from Site")
        else:
            indice = inv_fdf.loc[inv_fdf["Version"] == "Receptos"].index.values[0]
            add_to_excel(indice,"05.02.10","Y", f"Receptos FDF for {investigator[0]}, {investigator[1]}", "N")
    if investigator[2]< pd.Timestamp("2020-01-01") and ((investigator[3]=="Present") or (investigator[3] > pd.Timestamp("2016-03-01"))):
        if inv_fdf.loc[inv_fdf["Version"] == "Celgene"].empty:
            add_to_excel(0,"05.02.10", "N", f"Missing Celgene FDF for {investigator[0]}, {investigator[1]}", "Y","Collect from Site")
        else:
            indice = inv_fdf.loc[ inv_fdf["Version"] == "Celgene"].index.values[0]
            add_to_excel(indice,"05.02.10","Y", f"Celgene FDF for {investigator[0]}, {investigator[1]}", "N")
    if (investigator[3]=="Present") or (investigator[3] > pd.Timestamp("2020-01-15")):
            if inv_fdf.loc[inv_fdf["Version"] == "BMS"].empty:
                add_to_excel(0,"05.02.10", "N", f"Missing BMS FDF for {investigator[0]}, {investigator[1]}", "Collect from Site")
            else:
                indice = inv_fdf.loc[ inv_fdf["Version"] == "BMS"].index.values[0]
                add_to_excel(indice,"05.02.10","Y", f"BMS FDF for {investigator[0]}, {investigator[1]}", "N")




#%%
#TODO, integrar esto a el wizzard de configuracion
#IB
SFR = pd.read_excel(filename, sheet_name='Site',header=0)
if protocol == "RPC01-3101":
    IB = {
        "06" : pd.Timestamp("2015-02-09"),
        "07" : pd.Timestamp("2015-05-29"),
        "08" : pd.Timestamp("2016-05-02"),
        "09" : pd.Timestamp("2017-06-12"),
        "10" : pd.Timestamp("2018-04-30"),
        "11" : pd.Timestamp("2019-04-29")
         }
    PA = {    
        "01" : pd.Timestamp("2015-03-30"),
        "02" : pd.Timestamp("2016-06-07"),
        "03" : pd.Timestamp("2017-06-07"),
        "04" : pd.Timestamp("2017-12-07"),
        "05" : pd.Timestamp("2018-05-29"),
        "06": pd.Timestamp("2018-11-16"),
        "07": pd.Timestamp("2019-05-20")
         }
else:
    IB = {
        "06" : pd.Timestamp("2015-02-09"),
        "07" : pd.Timestamp("2015-05-29"),
        "08" : pd.Timestamp("2016-05-02"),
        "09" : pd.Timestamp("2017-06-12"),
        "10" : pd.Timestamp("2018-04-30"),
        "11" : pd.Timestamp("2019-04-26"),
        "12" : pd.Timestamp("2020-10-01")
         }
    PA = {    
        "01" : pd.Timestamp("2015-04-20"),
        "02" : pd.Timestamp("2015-05-08"),
        "03" : pd.Timestamp("2016-06-07"),
        "04" : pd.Timestamp("2017-06-07"),
        "05" : pd.Timestamp("2018-05-29"),
        "06": pd.Timestamp("2018-12-05"),
        "07": pd.Timestamp("2019-05-22"),
        "08": pd.Timestamp("2020-10-01")
         }
    
    
#IB
regex = re.compile(r'((\d)?\d)')
sp = SFR.loc[(SFR["Ref Model ID"] ==  "05.02.01") & (~(SFR["Document Name"]).isna())]
sp["Version"] = pd.Series(dtype="object")
for index,row in sp.iterrows():
    try:
        version = regex.search(sp["Document Name"][index]).group()
        
        if len(version) == 1:
            version = "0"+version
        sp["Version"][index] = version
    except:
        sp["Version"][index] = "Check this document"
        

ib_applicable = []
#re leo el visit_report, por las dudas
Visit_Report= pd.read_excel( os.getcwd()+"\\VISIT.xlsx",header=2)
Visit_Report = Visit_Report.loc[(Visit_Report["Protocol #"] == protocol) & (Visit_Report["Site #"] == Numero_de_sitio)]
spam = Visit_Report.loc[Visit_Report["Visit Type"] == "Site Visit Initiation"]
spam.reset_index(inplace=True)
start_date = spam["Visit Start"][0]
egg = Visit_Report.loc[(Visit_Report['Site #'] == Numero_de_sitio) & ((Visit_Report["Visit Type"] == "Site Visit Closeout") | (Visit_Report["Visit Type"] == "Telephone Closeout"))]
try:
    egg.reset_index(inplace=True)
    end_date = egg["Visit Start"][0]
except:
    end_date = datetime.datetime.today()
#para determinar con que IB empieza:
periodo_gracia = datetime.timedelta(days=60)
for key, value in IB.items():        
    if (start_date - value)> datetime.timedelta(days=0):       
        starting_ib = key
    if ((value - start_date) > datetime.timedelta(days=0)) and ((end_date - value) > periodo_gracia):
        ib_applicable.append(key)
        
if starting_ib not in ib_applicable:
    ib_applicable.insert(0,starting_ib)
    
#Copio la lista de ib aplicables para despues ponerlo en la parte de IRB
irb_ib_approvals = ib_applicable.copy()
#Agrego al excel la descripcion de los IBSP encontrados
ib_agregados = []
for index,row in sp.iterrows():
    version = str(sp['Version'][index])
    if version in ib_applicable:
        ib_applicable.remove(version)
    if version in ib_agregados:
        add_to_excel(index,"05.02.01", "Y", f"Potential duplicate for IB v{version} Signature Page", "Y", "Check if this is a duplicate")
        
    else:
        add_to_excel(index,"05.02.01", "Y", f"IB v{version} Signature Page", "N")
        ib_agregados.append(version)
    

if ib_applicable:
    for ibs in ib_applicable:
        add_to_excel(index,"05.02.01", "N", f"Missing IB v{ibs} Signature Page", "Y", "Collect from site")
        
#PA

SFR = pd.read_excel(filename, sheet_name='Site',header=0)
regex = re.compile(r'((\d)?\d)')
sp = SFR.loc[((SFR["Ref Model ID"] ==  "05.02.03") | (SFR["Ref Model ID"] ==  "05.02.02")) & (~(SFR["Document Name"]).isna())]
sp["Version"] = pd.Series(dtype="object")
for index,row in sp.iterrows():
    try:
        version = regex.search(sp["Document Name"][index]).group()        
        if len(version) == 1:
            version = "0"+version
        sp["Version"][index] = version
    except:
        sp["Version"][index] = "Check this document"
        

pa_applicable = []
#para determinar con que PA empieza:
periodo_gracia = datetime.timedelta(days=60)
for key, value in PA.items():        
    if (start_date - value)> datetime.timedelta(days=0):       
        starting_pa = key
    if ((value - start_date) > datetime.timedelta(days=0)) and ((end_date - value) > periodo_gracia):
        pa_applicable.append(key)
        
if starting_pa not in pa_applicable:
    pa_applicable.insert(0,starting_pa)
#Copio la lista de ib aplicables para despues ponerlo en la parte de IRB
irb_pa_approvals = pa_applicable.copy()
#Agrego al excel la descripcion de los PA SP encontrados
pa_agregados = []
for index,row in sp.iterrows():
    version = str(sp['Version'][index])
    if version == "01":
        code = "05.02.02"
    else:
        code = "05.02.03"
    if version in pa_applicable:
        pa_applicable.remove(version)
    if version in pa_agregados:
        add_to_excel(index,code, "Y", f"Potential duplicate for PA v{version} Signature Page", "Y", "Check if this is a duplicate")
        
    else:
        add_to_excel(index,code, "Y", f"PA v{version} Signature Page", "N")
        pa_agregados.append(version)   

if pa_applicable:
    for pas in pa_applicable:
        add_to_excel(index,code, "N", f"Missing PA v{pas} Signature Page", "Y", "Collect from site")
#%%
#Pequeño IRB reminder
if irb == "LIRB":
    msg = f"Please check that IRB approvals for PA {', '.join(irb_pa_approvals)} and IB {', '.join(irb_ib_approvals)} are present."
    add_to_excel(0,"04.01.02", "N",msg, "Y", "Check")
    #%%
    #IRB membership list
    SFR= pd.read_excel(filename, sheet_name='Site',header=0)
    SFR['Document date']=pd.to_datetime(SFR['Document date'])

    SFR_irb = SFR.loc[(SFR['Ref Model ID'] == '04.01.03') & ~(SFR['Document date'].isna() )]
    if SFR_irb.empty:
        add_to_excel(0, '04.01.03', 'N', f"IRB membership list missing. Please provide an updated version.", "N")
    else:
        SFR_irb.sort_values(by='Document date', inplace=True)
        SFR_irb=SFR_irb.tail(1)
        SFR_irb.reset_index(inplace=True)
        if (datetime.datetime.today() - SFR_irb['Document date'][0]) > datetime.timedelta(days=730):
            add_to_excel(0, '04.01.03', 'N', f"Last IRB membership list is from {SFR_irb['Document date'][0].date()}. Please provide an updated version.", "N")
#%%
#SDIL
SFR= pd.read_excel(filename, sheet_name='Site',header=0)
SFR['Document date']=pd.to_datetime(SFR['Document date'])

SFR_sdil = SFR.loc[(SFR['Ref Model ID'] == '05.04.02') & ~(SFR['Document date'].isna() )]
if SFR_sdil.empty:
    add_to_excel(0, '05.04.02', 'N', f"Missing Source Data Verification Log", "N")
else:
    add_to_excel(SFR_sdil.index.values[0], '05.04.02', 'Y', f"Source Data Verification Log.", "N")




#TODO
#Diferenciar entre smv y cov (con un input), y local/central irb(con un input tmb)
#en la parte de los logs, si tiene ip/sujetos pedir las cosas correspondientes, y si no no

#%%
#mega TODO
# import json
# if 'study_data.json' not in os.listdir('.'):
#     Study_data = {}
#     while True:
#         print('Introduzca lo que quiere agregar')
#         print('1.- Protocol Amendment\n2.- Investigator Brochure')
#         mode = input()
#         if mode == '':
#             break
#         while True:
            
#             print('Escriba por favor la informacion en el formato "Version-Fecha, siendo la fecha escriba en formato YYYYMMDD"')
#             print('Por ejemplo: "08-20200518"')
#             print('Si desea volver al otro menu, presione enter')
#             data = input()
#             if data == '':
#                 break
#             version = data.split('-')[0]
#             fecha = data.split('-')[1]
#             if len(fecha) != 8:
#                 print(f'La fecha introducida {fecha} parece que no esta en el formato YYYYMMDD. Por favor, re introducela')
#                 fecha = input()
#             if mode == 1:
#                 Study_data['PA V'+version] = fecha
#             else:
#                 Study_data['IB V'+version] = fecha
#             with open('study_data.json', 'w') as json_file:
#                 json.dump(Study_data, json_file)

#             print()
#             print('Si desea introducir otro amendment')
#         print('Si desea terminar, presione enter')
#%%
#FDA1572 estoy dormido, si funciona funciona
SFR= pd.read_excel(filename, sheet_name='Site',header=0)
SFR['Document date']=pd.to_datetime(SFR['Document date'])

SFR_FDA = SFR.loc[(SFR['Ref Model ID'] == '05.02.08') & ~(SFR['Document date'].isna() )]
SFR_FDA.sort_values(by='Document date', inplace=True)
SFR_FDA.reset_index(inplace=True)
if (SFR_FDA['Document date'][0] - Sitio.Site_Visit_Initiation[0]) > datetime.timedelta(days=365):
    add_to_excel(SFR_FDA['index'][0], '05.02.08', 'N', f"First FDA1572 is from {SFR_FDA['Document date'][0].date()} but site had its Site Visit Initiation in {Sitio.Site_Visit_Initiation[0]}. Please check",'N')
SFR_FDA=SFR_FDA.tail(1)
SFR_FDA.reset_index(inplace=True)
if Sitio.Cerrado == True:
    if hasattr(Sitio,'Site_Visit_Closeout'):
        if (datetime.datetime.strptime(Sitio.Site_Visit_Closeout[0], '%d-%b-%Y') - SFR_FDA['Document date'][0]) > datetime.timedelta(days=365):
            add_to_excel(SFR_FDA['index'][0], '05.02.08', 'N', f"Last FDA1572 is from {SFR_FDA['Document date'][0].date()} but site had its Site Visit Closeout in {Sitio.Site_Visit_Closeout[0]}. Please check if this is the latest one applicable",'N')
    else:
        if (datetime.datetime.strptime(Sitio.Telephone_Closeout[0], '%d-%b-%Y') - SFR_FDA['Document date'][0]) > datetime.timedelta(days=365):
            add_to_excel(SFR_FDA['index'][0], '05.02.08', 'N', f"Last FDA1572 is from {SFR_FDA['Document date'][0].date()} but site had its Telephone Closeout in {Sitio.Telephone_Closeout[0]}. Please check if this is the latest one applicable", "N")        
elif (datetime.datetime.today() - SFR_FDA['Document date'][0]) > datetime.timedelta(days=365):
    add_to_excel(SFR_FDA['index'][0], '05.02.08', 'N', f"Last FDA1572 is from {SFR_FDA['Document date'][0].date()}. Please check if we are not missing a more updated version.", "N")
